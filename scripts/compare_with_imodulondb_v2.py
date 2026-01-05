#!/usr/bin/env python3
"""
Compare gene review with iModulonDB data - Version 2 with CSV support.

Supports both Excel (P. putida) and CSV (E. coli) formats.

Usage:
    python scripts/compare_with_imodulondb_v2.py PSEPK BenR
    python scripts/compare_with_imodulondb_v2.py ecoli FliA
"""

import argparse
import sys
from pathlib import Path
from typing import Dict, List, Optional
import yaml
import openpyxl
import pandas as pd
from dataclasses import dataclass
import urllib.request


@dataclass
class OrganismInfo:
    """Information about organism's iModulonDB dataset."""
    taxon_id: str
    taxon_label: str
    imodulondb_code: str
    dataset: str
    github_repo: str
    github_branch: str
    supplementary_path: str
    reference: str
    doi: str
    available: bool
    data_format: Optional[str] = "excel"  # "excel" or "csv"


@dataclass
class IModulonMetrics:
    """Metrics for iModulon."""
    name: str
    regulator: Optional[str]
    imodulon_size: int
    precision: Optional[float] = None
    recall: Optional[float] = None
    f1_score: Optional[float] = None
    true_positives: Optional[int] = None
    regulon_size: Optional[int] = None
    pvalue: Optional[float] = None
    qvalue: Optional[float] = None
    category: Optional[str] = None
    subcategory: Optional[str] = None
    function: Optional[str] = None


@dataclass
class GeneWeight:
    """Gene with weight in iModulon."""
    locus_tag: str
    gene_name: Optional[str]
    weight: float
    product: Optional[str]


class IModulonDBClient:
    """Client for accessing iModulonDB data (Excel and CSV formats)."""

    def __init__(self, cache_dir: Path = Path(".cache/imodulondb")):
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.organisms = self._load_organism_mapping()

    def _load_organism_mapping(self) -> Dict[str, OrganismInfo]:
        """Load organism mapping from YAML file."""
        mapping_file = Path(__file__).parent.parent / "src/ai_gene_review/data/imodulondb_organisms.yaml"
        if not mapping_file.exists():
            print(f"Warning: Organism mapping file not found: {mapping_file}")
            return {}

        with open(mapping_file) as f:
            data = yaml.safe_load(f)

        organisms = {}
        for item in data.get("mappings", []):
            taxon_id = item["taxon_id"]
            organisms[taxon_id] = OrganismInfo(**item)

        return organisms

    def get_organism_by_code(self, org_code: str) -> Optional[OrganismInfo]:
        """Get organism info by short code."""
        code_to_taxon = {
            "PSEPK": "NCBITaxon:160488",
            "human": "NCBITaxon:9606",
            "ecoli": "NCBITaxon:511145",
        }
        taxon_id = code_to_taxon.get(org_code)
        return self.organisms.get(taxon_id) if taxon_id else None

    def download_csv_files(self, organism: OrganismInfo) -> Path:
        """Download CSV files for organisms like E. coli."""
        org_cache_dir = self.cache_dir / organism.imodulondb_code
        org_cache_dir.mkdir(parents=True, exist_ok=True)

        # Files to download
        files = [
            "imodulon_gene_names.txt",
            "TRN.csv",
            "gene_info.csv"
        ]

        for filename in files:
            cache_file = org_cache_dir / filename
            if not cache_file.exists():
                url = f"https://github.com/{organism.github_repo}/raw/{organism.github_branch}/{organism.supplementary_path}/{filename}"
                print(f"  Downloading {filename}...")
                try:
                    urllib.request.urlretrieve(url, cache_file)
                except Exception as e:
                    print(f"  Warning: Could not download {filename}: {e}")

        return org_cache_dir

    def parse_csv_dataset(self, cache_dir: Path, regulator: str) -> tuple:
        """Parse E. coli style CSV dataset."""
        # Read iModulon gene names
        imodulon_file = cache_dir / "imodulon_gene_names.txt"
        if not imodulon_file.exists():
            return None, []

        # Find the iModulon
        imodulon_genes = []
        imodulon_name = None

        with open(imodulon_file) as f:
            f.readline()  # Skip header
            for line in f:
                parts = line.strip().split(',')
                if parts[0] == regulator or regulator in parts[0]:
                    imodulon_name = parts[0]
                    imodulon_genes = [g for g in parts[1:] if g]
                    break

        if not imodulon_name:
            return None, []

        # Read gene info
        gene_info_file = cache_dir / "gene_info.csv"
        gene_info = {}
        if gene_info_file.exists():
            df = pd.read_csv(gene_info_file, index_col=0)
            for idx, row in df.iterrows():
                gene_info[row.get('gene_name', idx)] = {
                    'locus_tag': idx,
                    'gene_name': row.get('gene_name', ''),
                    'product': row.get('cog', 'Unknown')
                }

        # Create metrics
        metrics = IModulonMetrics(
            name=imodulon_name,
            regulator=regulator,
            imodulon_size=len(imodulon_genes)
        )

        # Create gene weights (CSV format doesn't have weights, use equal weight)
        gene_weights = []
        for i, gene_name in enumerate(imodulon_genes):
            info = gene_info.get(gene_name, {})
            gene_weights.append(GeneWeight(
                locus_tag=info.get('locus_tag', gene_name),
                gene_name=gene_name,
                weight=1.0 / (i + 1),  # Approximate weight by position
                product=info.get('product', 'Unknown')
            ))

        return metrics, gene_weights

    def download_supplementary_data(self, organism: OrganismInfo) -> Path:
        """Download supplementary Excel file from GitHub."""
        cache_file = self.cache_dir / organism.imodulondb_code / "supplementary_data.xlsx"

        if cache_file.exists():
            print(f"✓ Using cached data: {cache_file}")
            return cache_file

        cache_file.parent.mkdir(parents=True, exist_ok=True)

        url = f"https://github.com/{organism.github_repo}/raw/{organism.github_branch}/{organism.supplementary_path}"
        print(f"Downloading from {url}...")

        urllib.request.urlretrieve(url, cache_file)
        print(f"✓ Downloaded to {cache_file}")

        return cache_file

    def parse_excel_imodulon_table(self, excel_path: Path) -> Dict[str, IModulonMetrics]:
        """Parse iModulon table from Excel file (P. putida style)."""
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb['4-iM table']

        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        imodulons = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            imodulon = IModulonMetrics(
                name=row[0],
                regulator=row[1],
                pvalue=float(row[2]) if row[2] else 0,
                qvalue=float(row[3]) if row[3] else 0,
                precision=float(row[4]) if row[4] else 0,
                recall=float(row[5]) if row[5] else 0,
                f1_score=float(row[6]) if row[6] else 0,
                true_positives=int(row[7]) if row[7] else 0,
                regulon_size=int(row[8]) if row[8] else 0,
                imodulon_size=int(row[9]) if row[9] else 0,
                category=row[15] if len(row) > 15 else "",
                subcategory=row[16] if len(row) > 16 else "",
                function=row[17] if len(row) > 17 else ""
            )
            imodulons[imodulon.name] = imodulon

        return imodulons

    def extract_excel_gene_weights(self, excel_path: Path, imodulon_index: int,
                                   threshold: float = 0.05) -> List[GeneWeight]:
        """Extract genes with weights for specific iModulon (Excel format)."""
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        # Get gene info
        ws_genes = wb['2-Gene table']
        gene_data = {}
        for row in ws_genes.iter_rows(min_row=2, values_only=True):
            gene_data[row[0]] = {
                'locus_tag': row[0],
                'gene_name': row[1],
                'product': row[7] if len(row) > 7 else None
            }

        # Get M matrix
        ws_m = wb['6-M']
        gene_weights = []

        for row in ws_m.iter_rows(min_row=2, values_only=True):
            locus_tag = row[0]
            weight = row[imodulon_index + 1]

            if weight and abs(float(weight)) > threshold:
                if locus_tag in gene_data:
                    gene_weights.append(GeneWeight(
                        locus_tag=locus_tag,
                        gene_name=gene_data[locus_tag]['gene_name'],
                        weight=float(weight),
                        product=gene_data[locus_tag]['product']
                    ))

        gene_weights.sort(key=lambda g: abs(g.weight), reverse=True)
        return gene_weights

    def find_imodulon_index(self, excel_path: Path, imodulon_name: str) -> Optional[int]:
        """Find the index of an iModulon in the Excel dataset."""
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb['4-iM table']

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if row[0] == imodulon_name:
                return i

        return None

    def compare_with_review(self, organism_code: str, gene_symbol: str,
                          output_dir: Path = Path(".")) -> None:
        """Compare gene review with iModulonDB data."""

        # 1. Get organism info
        organism = self.get_organism_by_code(organism_code)
        if not organism:
            print(f"❌ Organism '{organism_code}' not found in iModulonDB mapping")
            return

        if not organism.available:
            print(f"❌ iModulonDB data not available for {organism.taxon_label}")
            return

        print(f"📊 Comparing {gene_symbol} with iModulonDB dataset: {organism.dataset}")

        # 2. Download and parse data based on format
        if organism.data_format == "csv":
            print(f"📖 Downloading CSV files...")
            cache_dir = self.download_csv_files(organism)
            print(f"📖 Parsing CSV dataset...")
            imodulon, gene_weights = self.parse_csv_dataset(cache_dir, gene_symbol)

            if not imodulon:
                print(f"⚠️  No iModulon found for regulator '{gene_symbol}'")
                return

            print(f"✓ Found {imodulon.name} iModulon ({len(gene_weights)} genes)")

        else:  # Excel format
            excel_path = self.download_supplementary_data(organism)
            print(f"📖 Parsing iModulon table...")
            imodulons = self.parse_excel_imodulon_table(excel_path)

            imodulon = None
            for name, metrics in imodulons.items():
                if metrics.regulator == gene_symbol or name == gene_symbol:
                    imodulon = metrics
                    break

            if not imodulon:
                print(f"⚠️  No iModulon found for regulator '{gene_symbol}'")
                return

            print(f"✓ Found {imodulon.name} iModulon")

            imodulon_index = self.find_imodulon_index(excel_path, imodulon.name)
            if imodulon_index is None:
                print(f"❌ Could not find iModulon index")
                return

            print(f"📊 Extracting gene weights...")
            gene_weights = self.extract_excel_gene_weights(excel_path, imodulon_index)

        # 3. Generate report
        self._generate_report(organism, gene_symbol, imodulon, gene_weights, output_dir)

    def _generate_report(self, organism: OrganismInfo, gene_symbol: str,
                        imodulon: IModulonMetrics, genes: List[GeneWeight],
                        output_dir: Path) -> None:
        """Generate comparison report."""

        report_file = output_dir / f"genes/{organism.imodulondb_code}/{gene_symbol}/{gene_symbol}-imodulondb-comparison.md"
        report_file.parent.mkdir(parents=True, exist_ok=True)

        with open(report_file, 'w') as f:
            f.write(f"# {gene_symbol} iModulonDB Comparison\n\n")

            f.write(f"## Dataset Information\n\n")
            f.write(f"- **Organism**: {organism.taxon_label}\n")
            f.write(f"- **Dataset**: {organism.dataset}\n")
            f.write(f"- **Reference**: {organism.reference}\n")
            f.write(f"- **DOI**: {organism.doi}\n\n")

            f.write(f"## {imodulon.name} iModulon Statistics\n\n")
            f.write(f"| Metric | Value |\n")
            f.write(f"|--------|-------|\n")
            f.write(f"| **iModulon Size** | {imodulon.imodulon_size} genes |\n")

            if imodulon.regulon_size:
                f.write(f"| **Known Regulon Size** | {imodulon.regulon_size} genes |\n")
                f.write(f"| **True Positives** | {imodulon.true_positives} |\n")
                f.write(f"| **Precision** | {imodulon.precision:.2f} ({imodulon.precision*100:.0f}%) |\n")
                f.write(f"| **Recall** | {imodulon.recall:.2f} ({imodulon.recall*100:.0f}%) |\n")
                f.write(f"| **F1 Score** | {imodulon.f1_score:.2f} |\n")

            if imodulon.category:
                f.write(f"| **Category** | {imodulon.category}")
                if imodulon.subcategory:
                    f.write(f" - {imodulon.subcategory}")
                f.write(" |\n")

            if imodulon.function:
                f.write(f"| **Function** | {imodulon.function} |\n")

            f.write(f"\n## Genes in iModulon\n\n")
            f.write(f"| Rank | Locus Tag | Gene | Weight | Product |\n")
            f.write(f"|------|-----------|------|--------|----------|\n")

            for i, gene in enumerate(genes[:20], 1):
                gene_name = gene.gene_name if gene.gene_name else "---"
                product = gene.product[:60] if gene.product else "Unknown"
                f.write(f"| {i} | {gene.locus_tag} | {gene_name} | {gene.weight:.4f} | {product} |\n")

            f.write(f"\n**Total genes in iModulon**: {len(genes)}\n\n")

            # Interpretation
            f.write(f"## Interpretation\n\n")

            if imodulon.recall and imodulon.f1_score:
                consistency = "HIGH" if imodulon.recall >= 0.8 and imodulon.f1_score >= 0.5 else \
                             "MEDIUM" if imodulon.recall >= 0.6 and imodulon.f1_score >= 0.4 else "LOW"
                f.write(f"**Consistency Level**: {consistency}\n\n")

                if imodulon.recall == 1.0:
                    f.write(f"✅ **Perfect Recall**: All known regulon members are captured in the iModulon.\n\n")

                if imodulon.precision and imodulon.precision < 0.5:
                    f.write(f"⚠️ **Moderate Precision**: The iModulon contains additional genes beyond ")
                    f.write(f"the validated regulon, suggesting functional coupling or novel predictions.\n\n")
            else:
                f.write(f"**Note**: CSV format datasets don't include precision/recall metrics. ")
                f.write(f"Gene ordering approximates importance.\n\n")

        print(f"✅ Report generated: {report_file}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Compare gene review with iModulonDB data",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("organism", help="Organism code (e.g., PSEPK, ecoli)")
    parser.add_argument("gene", help="Gene symbol (e.g., BenR, FliA)")
    parser.add_argument("--cache-dir", default=".cache/imodulondb",
                       help="Cache directory for downloaded data")
    parser.add_argument("--output-dir", default=".",
                       help="Output directory for comparison report")

    args = parser.parse_args()

    client = IModulonDBClient(cache_dir=Path(args.cache_dir))
    client.compare_with_review(
        organism_code=args.organism,
        gene_symbol=args.gene,
        output_dir=Path(args.output_dir)
    )


if __name__ == "__main__":
    main()
