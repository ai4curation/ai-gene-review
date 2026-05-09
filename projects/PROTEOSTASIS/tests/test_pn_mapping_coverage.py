"""Tests for PN mapping coverage reporting."""

from pathlib import Path

from openpyxl import Workbook
import yaml

from ai_gene_review.tools.report_pn_mapping_coverage import (
    load_subject_specs,
    load_workbook_codes,
    summarize_coverage,
)


def _write_minimal_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "dense"
    sheet.append(["Gene Symbol", "Branch", "Class", "Group", "Type", "Subtype"])
    sheet.append(
        [
            "GENE1",
            "Translation",
            "Cytosolic translation",
            "Translation initiation",
            "eIF2 complex",
            None,
        ]
    )
    sheet.append(
        [
            "GENE2",
            "Translation",
            "Cytosolic translation",
            "Translation initiation",
            "eIF4F complex",
            None,
        ]
    )
    sheet.append(
        [
            "GENE3",
            "Translation",
            "Cytosolic translation",
            "Translation initiation",
            "assorted initiation factors",
            None,
        ]
    )
    workbook.save(path)


def test_coverage_distinguishes_curation_statuses_and_missing_yaml(tmp_path: Path) -> None:
    workbook_path = tmp_path / "mini_pn.xlsx"
    _write_minimal_workbook(workbook_path)

    mapping_dir = tmp_path / "mappings"
    mapping_dir.mkdir()
    mapping_set = {
        "id": "pn_mapping_set:test",
        "label": "Test mapping set",
        "references": ["proteostasis-workbook-2024"],
        "subject_curations": [
            {
                "subject_code": "Translation|Cytosolic translation|Translation initiation|eIF2 complex",
                "subject_level": "type",
                "curation_status": "mapped",
                "target_term": {
                    "id": "GO:0005850",
                    "label": "eukaryotic translation initiation factor 2 complex",
                },
                "mapping_scope": "ok_for_propagation_to_go",
                "rationale": "Test mapped entry.",
            },
            {
                "subject_code": "Translation|Cytosolic translation|Translation initiation|eIF4F complex",
                "subject_level": "type",
                "curation_status": "no_mapping",
                "rationale": "Test explicit unmapped entry.",
            }
        ],
    }
    (mapping_dir / "translation.yaml").write_text(
        yaml.safe_dump(mapping_set, sort_keys=False),
        encoding="utf-8",
    )

    code_records = load_workbook_codes(workbook_path)
    curation_specs = load_subject_specs(mapping_dir)
    coverage_rows = summarize_coverage(code_records, curation_specs)

    status_by_code = {
        row["code"]: row["status"] for row in coverage_rows if row["level"] == "type"
    }
    assert (
        status_by_code["Translation|Cytosolic translation|Translation initiation|eIF2 complex"]
        == "mapped"
    )
    assert (
        status_by_code["Translation|Cytosolic translation|Translation initiation|eIF4F complex"]
        == "no_mapping"
    )
    assert (
        status_by_code[
            "Translation|Cytosolic translation|Translation initiation|assorted initiation factors"
        ]
        == "missing_from_yaml"
    )
