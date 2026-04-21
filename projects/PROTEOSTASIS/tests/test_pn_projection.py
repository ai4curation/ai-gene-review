"""Tests for PN-driven GO projection reporting."""

from pathlib import Path

from openpyxl import Workbook
import yaml

from ai_gene_review.tools.report_pn_projected_annotations import (
    load_mapping_specs,
    load_workbook_rows,
    project_annotations,
    summarize_gene_go_projections,
    write_projection_reports,
)


def _write_projection_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "dense"
    sheet.append(
        [
            "order",
            "Gene Symbol",
            "Gene Name",
            "Branch",
            "Class",
            "Group",
            "Type",
            "Subtype",
            "UniProt ID",
        ]
    )
    sheet.append(
        [
            1,
            "GENE1",
            "Gene one",
            "Cytonuclear proteostasis",
            "Chaperone",
            "HSP70 system",
            "HSP70",
            None,
            "P00001",
        ]
    )
    sheet.append(
        [
            2,
            "GENE2",
            "Gene two",
            "Cytonuclear proteostasis",
            "Chaperone",
            "HSP70 system",
            "HSP70",
            None,
            "P00002",
        ]
    )
    sheet.append(
        [
            3,
            "GENE3",
            "Gene three",
            "Cytonuclear proteostasis",
            "Chaperone",
            None,
            None,
            None,
            "P00003",
        ]
    )
    sheet.append(
        [
            4,
            "GENE4",
            "Gene four",
            "Cytonuclear proteostasis",
            "Chaperone",
            "HSP70 system",
            "HSP70",
            None,
            "P00004",
        ]
    )
    sheet.append(
        [
            5,
            "GENE5",
            "Gene five",
            "Cytonuclear proteostasis",
            "Chaperone",
            "HSP70 system",
            "HSP70",
            None,
            "P00005",
        ]
    )
    workbook.save(path)


def _goa_tsv_text(gene_product_id: str, gene_symbol: str, go_id: str, go_label: str) -> str:
    return "\n".join(
        [
            "\t".join(
                [
                    "DB",
                    "GENE PRODUCT ID",
                    "SYMBOL",
                    "QUALIFIER",
                    "GO TERM",
                    "GO NAME",
                    "GO ASPECT",
                    "ECO ID",
                    "GO EVIDENCE CODE",
                    "REFERENCE",
                    "WITH/FROM",
                    "TAXON ID",
                    "TAXON NAME",
                    "ASSIGNED BY",
                    "GENE NAME",
                    "DATE",
                ]
            ),
            "\t".join(
                [
                    "UniProtKB",
                    gene_product_id,
                    gene_symbol,
                    "",
                    go_id,
                    go_label,
                    "C",
                    "ECO:0000314",
                    "IDA",
                    "PMID:1",
                    "",
                    "taxon:9606",
                    "Homo sapiens",
                    "UniProt",
                    gene_symbol,
                    "20260101",
                ]
            ),
        ]
    )


def _write_minimal_goa(path: Path, gene_symbol: str, go_id: str, go_label: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(_goa_tsv_text("P00001", gene_symbol, go_id, go_label), encoding="utf-8")


def test_pn_projection_reports_existing_new_and_missing_goa(tmp_path: Path) -> None:
    workbook_path = tmp_path / "pn.xlsx"
    _write_projection_workbook(workbook_path)

    mapping_dir = tmp_path / "mappings"
    mapping_dir.mkdir()
    mapping_set = {
        "id": "pn_mapping_set:test",
        "label": "Test mapping set",
        "mappings": [
            {
                "subject_code": "Cytonuclear proteostasis|Chaperone",
                "subject_level": "class",
                "target_term": {
                    "id": "GO:0044183",
                    "label": "protein folding chaperone",
                },
                "mapping_scope": "ok_for_propagation_to_go",
                "representative_genes": ["HSPA1A", "DNAJB1"],
                "rationale": "Broad chaperone mapping.",
                "references": ["proteostasis-ms1"],
            },
            {
                "subject_code": "HSP70",
                "subject_level": "type",
                "target_term": {
                    "id": "GO:0140662",
                    "label": "ATP-dependent protein folding chaperone",
                },
                "mapping_scope": "ok_for_propagation_to_go",
                "representative_genes": ["HSPA1A", "HSPA8"],
                "conditions": [
                    {"condition_level": "class", "condition_code": "Chaperone"},
                    {"condition_level": "group", "condition_code": "HSP70 system"},
                ],
                "rationale": "Test projection mapping.",
                "references": ["proteostasis-ms1"],
            },
        ],
    }
    (mapping_dir / "translation.yaml").write_text(
        yaml.safe_dump(mapping_set, sort_keys=False),
        encoding="utf-8",
    )

    goa_root = tmp_path / "genes" / "human"
    _write_minimal_goa(
        goa_root / "GENE1" / "GENE1-goa.tsv",
        gene_symbol="GENE1",
        go_id="GO:0140662",
        go_label="ATP-dependent protein folding chaperone",
    )
    _write_minimal_goa(
        goa_root / "GENE2" / "GENE2-goa.tsv",
        gene_symbol="GENE2",
        go_id="GO:0044183",
        go_label="protein folding chaperone",
    )
    _write_minimal_goa(
        goa_root / "GENE3" / "GENE3-goa.tsv",
        gene_symbol="GENE3",
        go_id="GO:0140662",
        go_label="ATP-dependent protein folding chaperone",
    )
    _write_minimal_goa(
        goa_root / "GENE5" / "GENE5-goa.tsv",
        gene_symbol="GENE5",
        go_id="GO:0016887",
        go_label="ATP hydrolysis activity",
    )

    workbook_rows = load_workbook_rows(workbook_path)
    mapping_specs = load_mapping_specs(mapping_dir)
    projected_rows = project_annotations(workbook_rows, mapping_specs, goa_root)
    summaries = summarize_gene_go_projections(projected_rows)

    status_by_gene_go = {
        (row["gene_symbol"], row["target_go_id"]): row["goa_status"] for row in summaries
    }
    assert status_by_gene_go[("GENE1", "GO:0140662")] == "already_in_goa_exact"
    assert status_by_gene_go[("GENE1", "GO:0044183")] == "entailed_by_goa_closure"
    assert status_by_gene_go[("GENE2", "GO:0140662")] == "more_specific_than_existing_goa"
    assert status_by_gene_go[("GENE2", "GO:0044183")] == "already_in_goa_exact"
    assert status_by_gene_go[("GENE3", "GO:0044183")] == "entailed_by_goa_closure"
    assert status_by_gene_go[("GENE4", "GO:0140662")] == "no_local_goa"
    assert status_by_gene_go[("GENE5", "GO:0140662")] == "new_to_goa"

    gene2_summary = next(
        row
        for row in summaries
        if row["gene_symbol"] == "GENE2" and row["target_go_id"] == "GO:0140662"
    )
    assert gene2_summary["representative_genes"] == "HSPA1A;HSPA8"
    assert gene2_summary["supporting_goa_terms"] == "GO:0044183 protein folding chaperone"

    output_dir = tmp_path / "reports"
    write_projection_reports(output_dir, projected_rows, summaries)
    assert (output_dir / "pn_projected_annotations.tsv").exists()
    assert (output_dir / "pn_projected_gene_go_summary.tsv").exists()
    assert (output_dir / "pn_projected_new_to_goa.tsv").exists()
    assert (output_dir / "pn_projected_candidate_additions.tsv").exists()
    assert (output_dir / "pn_projected_status_summary.tsv").exists()


def test_pn_projection_can_fetch_missing_goa_into_project_cache(
    tmp_path: Path,
    monkeypatch,
) -> None:
    workbook_path = tmp_path / "pn.xlsx"
    _write_projection_workbook(workbook_path)

    mapping_dir = tmp_path / "mappings"
    mapping_dir.mkdir()
    mapping_set = {
        "id": "pn_mapping_set:test",
        "label": "Test mapping set",
        "mappings": [
            {
                "subject_code": "HSP70",
                "subject_level": "type",
                "target_term": {
                    "id": "GO:0140662",
                    "label": "ATP-dependent protein folding chaperone",
                },
                "mapping_scope": "ok_for_propagation_to_go",
                "rationale": "Test projection mapping.",
                "conditions": [
                    {"condition_level": "class", "condition_code": "Chaperone"},
                    {"condition_level": "group", "condition_code": "HSP70 system"},
                ],
                "references": ["proteostasis-ms1"],
            }
        ],
    }
    (mapping_dir / "chaperone_systems.yaml").write_text(
        yaml.safe_dump(mapping_set, sort_keys=False),
        encoding="utf-8",
    )

    goa_root = tmp_path / "genes" / "human"
    cache_dir = tmp_path / "cache" / "goa"

    def fake_fetch_goa_data(uniprot_id: str) -> str:
        assert uniprot_id == "P00004"
        return _goa_tsv_text(
            "P00004",
            "GENE4",
            "GO:0140662",
            "ATP-dependent protein folding chaperone",
        )

    monkeypatch.setattr(
        "ai_gene_review.tools.report_pn_projected_annotations.fetch_goa_data",
        fake_fetch_goa_data,
    )

    workbook_rows = load_workbook_rows(workbook_path)
    mapping_specs = load_mapping_specs(mapping_dir)
    projected_rows = project_annotations(
        [row for row in workbook_rows if row.gene_symbol == "GENE4"],
        mapping_specs,
        goa_root,
        fetch_missing_goa=True,
        goa_cache_dir=cache_dir,
    )
    summaries = summarize_gene_go_projections(projected_rows)

    assert summaries[0]["goa_status"] == "already_in_goa_exact"
    assert (cache_dir / "GENE4__P00004-goa.tsv").exists()
    assert not (goa_root / "GENE4" / "GENE4-goa.tsv").exists()
