"""Tests for XLSX export of external mapping YAML files."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl import Workbook as OpenPyxlWorkbook
import yaml

from ai_gene_review.tools.export_ext_mappings_xlsx import (
    apply_derived_representative_genes,
    build_projection_tables,
    build_mapping_row_representative_lookup,
    build_workbook,
    load_mapping_rows,
)


def test_export_builds_biologist_friendly_workbook(tmp_path: Path) -> None:
    mapping_dir = tmp_path / "mappings"
    mapping_dir.mkdir()
    mapping_set = {
        "id": "pn_mapping_set:test",
        "label": "Test mapping set",
        "references": ["proteostasis-workbook-2024"],
        "mappings": [
            {
                "subject_code": "Translation|Cytosolic translation|Ribosome",
                "subject_level": "group",
                "target_term": {
                    "id": "GO:0022626",
                    "label": "cytosolic ribosome",
                },
                "mapping_scope": "ok_for_propagation_to_go",
                "representative_genes": ["EIF2S1", "EIF2S2"],
                "conditions": [
                    {"condition_level": "branch", "condition_code": "Translation"}
                ],
                "rationale": "Test mapping.",
                "notes": "Test note.",
                "references": ["proteostasis-ms1"],
            }
        ],
        "unmapped_subjects": [
            {
                "subject_code": "Translation|Mitochondrial translation|Ribosome",
                "subject_level": "group",
                "rationale": "No term available in current cache.",
                "references": ["proteostasis-ms1"],
            }
        ],
    }
    path = mapping_dir / "translation.yaml"
    path.write_text(yaml.safe_dump(mapping_set, sort_keys=False), encoding="utf-8")

    mapping_rows, unmapped_rows = load_mapping_rows(mapping_dir)
    workbook = build_workbook(mapping_rows, unmapped_rows)

    output_path = tmp_path / "pn_mappings.xlsx"
    workbook.save(output_path)
    reopened = load_workbook(output_path)

    assert reopened.sheetnames[:5] == [
        "README",
        "Summary",
        "Mappings",
        "Unmapped",
        "Translation",
    ]

    readme = reopened["README"]
    assert readme["A1"].value == "Proteostasis PN to GO mapping export"
    assert "manually curated mappings" in readme["B2"].value
    assert "pipe-delimited PN path" in readme["B3"].value

    summary = reopened["Summary"]
    assert summary["A2"].value == "translation.yaml"
    assert summary["D2"].value == "1"
    assert summary["E2"].value == "1"

    mappings = reopened["Mappings"]
    assert mappings["E2"].value == "Translation|Cytosolic translation|Ribosome"
    assert mappings["F2"].value == "Translation"
    assert mappings["G2"].value == "Cytosolic translation"
    assert mappings["H2"].value == "Ribosome"
    assert mappings["L2"].value == "GO:0022626"
    assert mappings["M2"].value == "cytosolic ribosome"
    assert mappings["N2"].value == "EIF2S1; EIF2S2"
    assert mappings["O2"].value == "branch=Translation"

    unmapped = reopened["Unmapped"]
    assert unmapped["E2"].value == "Translation|Mitochondrial translation|Ribosome"
    assert unmapped["F2"].value == "Translation"
    assert unmapped["G2"].value == "Mitochondrial translation"
    assert unmapped["H2"].value == "Ribosome"

    per_file = reopened["Translation"]
    assert per_file["A2"].value == "mapped"
    assert per_file["A3"].value == "explicit_unmapped"
    assert per_file["L2"].value == "EIF2S1; EIF2S2"


def test_export_can_embed_projection_tabs(tmp_path: Path) -> None:
    mapping_dir = tmp_path / "mappings"
    mapping_dir.mkdir()
    mapping_set = {
        "id": "pn_mapping_set:test",
        "label": "Test mapping set",
        "references": ["proteostasis-workbook-2024"],
        "mappings": [
            {
                "subject_code": "Cytonuclear proteostasis|Chaperone",
                "subject_level": "class",
                "target_term": {
                    "id": "GO:0044183",
                    "label": "protein folding chaperone",
                },
                "mapping_scope": "ok_for_propagation_to_go",
                "representative_genes": ["HSPA1A", "DNAJB1"],
                "rationale": "Broad chaperone mapping.",
                "references": ["proteostasis-ms1"],
            },
            {
                "subject_code": "HSP70",
                "subject_level": "type",
                "target_term": {
                    "id": "GO:0140662",
                    "label": "ATP-dependent protein folding chaperone",
                },
                "mapping_scope": "ok_for_propagation_to_go",
                "representative_genes": ["HSPA1A", "HSPA8"],
                "conditions": [
                    {"condition_level": "class", "condition_code": "Chaperone"},
                    {"condition_level": "group", "condition_code": "HSP70 system"},
                ],
                "rationale": "Specific HSP70 mapping.",
                "references": ["proteostasis-ms1"],
            },
        ],
    }
    (mapping_dir / "chaperone_systems.yaml").write_text(
        yaml.safe_dump(mapping_set, sort_keys=False),
        encoding="utf-8",
    )

    workbook_path = tmp_path / "pn.xlsx"
    wb = OpenPyxlWorkbook()
    sheet = wb.active
    sheet.title = "dense"
    sheet.append(
        [
            "order",
            "Gene Symbol",
            "Gene Name",
            "Branch",
            "Class",
            "Group",
            "Type",
            "Subtype",
        ]
    )
    sheet.append([1, "GENE1", "Gene one", "Cytonuclear proteostasis", "Chaperone", "HSP70 system", "HSP70", None])
    sheet.append([2, "GENE2", "Gene two", "Cytonuclear proteostasis", "Chaperone", "HSP70 system", "HSP70", None])
    wb.save(workbook_path)

    goa_root = tmp_path / "genes" / "human"
    gene1_dir = goa_root / "GENE1"
    gene1_dir.mkdir(parents=True)
    (gene1_dir / "GENE1-goa.tsv").write_text(
        "\n".join(
            [
                "\t".join(
                    [
                        "DB",
                        "GENE PRODUCT ID",
                        "SYMBOL",
                        "QUALIFIER",
                        "GO TERM",
                        "GO NAME",
                        "GO ASPECT",
                        "ECO ID",
                        "GO EVIDENCE CODE",
                        "REFERENCE",
                        "WITH/FROM",
                        "TAXON ID",
                        "TAXON NAME",
                        "ASSIGNED BY",
                        "GENE NAME",
                        "DATE",
                    ]
                ),
                "\t".join(
                    [
                        "UniProtKB",
                        "P00001",
                        "GENE1",
                        "",
                        "GO:0044183",
                        "protein folding chaperone",
                        "molecular_function",
                        "ECO:0000314",
                        "IDA",
                        "PMID:1",
                        "",
                        "taxon:9606",
                        "Homo sapiens",
                        "UniProt",
                        "GENE1",
                        "20260101",
                    ]
                ),
            ]
        ),
        encoding="utf-8",
    )

    mapping_rows, unmapped_rows = load_mapping_rows(mapping_dir)
    projected_summary_rows, candidate_rows, projection_status_rows = build_projection_tables(
        mapping_dir,
        workbook_path,
        goa_root,
    )
    workbook = build_workbook(
        mapping_rows,
        unmapped_rows,
        projected_summary_rows=projected_summary_rows,
        candidate_rows=candidate_rows,
        projection_status_rows=projection_status_rows,
    )

    output_path = tmp_path / "pn_mappings_with_projection.xlsx"
    workbook.save(output_path)
    reopened = load_workbook(output_path)

    assert "Projected Gene-GO" in reopened.sheetnames
    assert "Candidate Additions" in reopened.sheetnames
    assert "Projection Status" in reopened.sheetnames

    projected = reopened["Projected Gene-GO"]
    assert projected["A2"].value == "GENE1"
    assert projected["E2"].value in {
        "already_in_goa_exact",
        "more_specific_than_existing_goa",
        "entailed_by_goa_closure",
    }

    candidate = reopened["Candidate Additions"]
    assert candidate["A2"].value == "GENE1" or candidate["A2"].value == "GENE2"

    status = reopened["Projection Status"]
    assert status["A2"].value == "row_level_projections"


def test_export_can_fill_blank_representative_genes_from_workbook(tmp_path: Path) -> None:
    mapping_dir = tmp_path / "mappings"
    mapping_dir.mkdir()
    mapping_set = {
        "id": "pn_mapping_set:test",
        "label": "Test mapping set",
        "references": ["proteostasis-workbook-2024"],
        "mappings": [
            {
                "subject_code": "Cytonuclear proteostasis|Chaperone",
                "subject_level": "class",
                "target_term": {
                    "id": "GO:0044183",
                    "label": "protein folding chaperone",
                },
                "mapping_scope": "ok_for_propagation_to_go",
                "rationale": "Broad chaperone mapping.",
                "references": ["proteostasis-ms1"],
            }
        ],
    }
    (mapping_dir / "chaperone_systems.yaml").write_text(
        yaml.safe_dump(mapping_set, sort_keys=False),
        encoding="utf-8",
    )

    workbook_path = tmp_path / "pn.xlsx"
    wb = OpenPyxlWorkbook()
    sheet = wb.active
    sheet.title = "dense"
    sheet.append(
        [
            "order",
            "Gene Symbol",
            "Gene Name",
            "Branch",
            "Class",
            "Group",
            "Type",
            "Subtype",
        ]
    )
    sheet.append([1, "GENE1", "Gene one", "Cytonuclear proteostasis", "Chaperone", "HSP70 system", "HSP70", None])
    sheet.append([2, "GENE2", "Gene two", "Cytonuclear proteostasis", "Chaperone", "HSP90 system", "HSP90", None])
    wb.save(workbook_path)

    mapping_rows, unmapped_rows = load_mapping_rows(mapping_dir)
    representative_lookup = build_mapping_row_representative_lookup(mapping_dir, workbook_path)
    mapping_rows = apply_derived_representative_genes(mapping_rows, representative_lookup)
    workbook = build_workbook(mapping_rows, unmapped_rows)

    output_path = tmp_path / "pn_mappings_with_examples.xlsx"
    workbook.save(output_path)
    reopened = load_workbook(output_path)

    mappings = reopened["Mappings"]
    assert mappings["N2"].value == "GENE1; GENE2"
