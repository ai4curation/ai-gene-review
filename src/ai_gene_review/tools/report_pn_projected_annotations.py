"""Project GO annotations from PN workbook rows through curated PN-to-GO mappings."""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import yaml

from ai_gene_review.etl.gene import fetch_goa_data
from ai_gene_review.tools.report_pn_mapping_coverage import LEVELS, WORKBOOK_COLUMNS
from ai_gene_review.validation.goa_validator import GOAValidator

DEFAULT_WORKBOOK_SHEET = "dense"
DEFAULT_MAPPING_DIR = Path("projects/PROTEOSTASIS/mappings")
DEFAULT_GOA_ROOT = Path("genes/human")
DEFAULT_OUTPUT_DIR = Path("projects/PROTEOSTASIS/reports/pn_projection")
DEFAULT_GOA_CACHE_DIR = Path("projects/PROTEOSTASIS/cache/goa")
DEFAULT_REPRESENTATIVE_GENE_LIMIT = 5
MISSING_MARKERS = {"", "(no Branch)", "(no Class)", "(no Group)", "(no Type)", "(no Subtype)"}
GO_CLOSURE_PREDICATES = ("rdfs:subClassOf", "BFO:0000050")


@dataclass(frozen=True)
class PNWorkbookRow:
    """One gene-level annotation row from the PN workbook."""

    order: int | None
    gene_symbol: str
    gene_name: str
    branch: str
    class_name: str
    group: str
    type_name: str
    subtype: str
    uniprot_id: str


@dataclass(frozen=True)
class MappingSpec:
    """One curated PN-to-GO mapping entry."""

    file_name: str
    subject_code: str
    subject_level: str
    target_go_id: str
    target_go_label: str
    mapping_scope: str
    representative_genes: tuple[str, ...]
    conditions: tuple[tuple[str, str], ...]
    rationale: str
    notes: str
    references: tuple[str, ...]


@dataclass(frozen=True)
class ProjectedAnnotation:
    """A GO term projected onto one PN workbook gene row."""

    gene_symbol: str
    gene_name: str
    pn_code: str
    branch: str
    class_name: str
    group: str
    type_name: str
    subtype: str
    mapping_file: str
    mapping_subject_level: str
    mapping_subject_code: str
    mapping_scope: str
    target_go_id: str
    target_go_label: str
    representative_genes: tuple[str, ...]
    conditions: tuple[tuple[str, str], ...]
    rationale: str
    notes: str
    references: tuple[str, ...]
    goa_status: str
    supporting_goa_terms: tuple[str, ...]


def _clean_value(value: Any) -> str:
    if value is None:
        return ""
    cleaned = str(value).strip()
    if cleaned in MISSING_MARKERS:
        return ""
    return cleaned


def _canonical_code(level_values: dict[str, str], level: str) -> str:
    parts: list[str] = []
    for current_level in LEVELS:
        value = level_values[current_level]
        if not value:
            break
        parts.append(value)
        if current_level == level:
            break
    return "|".join(parts)


def _normalize_conditions(raw_conditions: Any) -> tuple[tuple[str, str], ...]:
    if not raw_conditions:
        return ()
    conditions: list[tuple[str, str]] = []
    for condition in raw_conditions:
        conditions.append(
            (
                _clean_value(condition.get("condition_level")),
                _clean_value(condition.get("condition_code")),
            )
        )
    return tuple(conditions)


def _level_value(row: PNWorkbookRow, level: str) -> str:
    if level == "branch":
        return row.branch
    if level == "class":
        return row.class_name
    if level == "group":
        return row.group
    if level == "type":
        return row.type_name
    if level == "subtype":
        return row.subtype
    raise ValueError(f"Unsupported level: {level}")


def _code_at_level(row: PNWorkbookRow, level: str) -> str:
    return _canonical_code(
        {
            "branch": row.branch,
            "class": row.class_name,
            "group": row.group,
            "type": row.type_name,
            "subtype": row.subtype,
        },
        level,
    )


def _deepest_code(row: PNWorkbookRow) -> str:
    for level in reversed(LEVELS):
        if _level_value(row, level):
            return _code_at_level(row, level)
    return ""


def _row_sort_key(row: PNWorkbookRow) -> tuple[int, str]:
    if row.order is None:
        return (10**9, row.gene_symbol)
    return (row.order, row.gene_symbol)


def _matches(row: PNWorkbookRow, spec: MappingSpec) -> bool:
    if _level_value(row, spec.subject_level) == "":
        return False

    if "|" in spec.subject_code:
        if _code_at_level(row, spec.subject_level) != spec.subject_code:
            return False
    elif _level_value(row, spec.subject_level) != spec.subject_code:
        return False

    for condition_level, condition_code in spec.conditions:
        if "|" in condition_code:
            if _code_at_level(row, condition_level) != condition_code:
                return False
        elif _level_value(row, condition_level) != condition_code:
            return False

    return True


def derive_representative_genes(
    workbook_rows: list[PNWorkbookRow],
    mapping_specs: list[MappingSpec],
    max_genes: int = DEFAULT_REPRESENTATIVE_GENE_LIMIT,
) -> dict[MappingSpec, tuple[str, ...]]:
    """Fill mapping exemplars from workbook membership when curation left them blank."""
    sorted_rows = sorted(workbook_rows, key=_row_sort_key)
    representative_genes_by_spec: dict[MappingSpec, tuple[str, ...]] = {}

    for spec in mapping_specs:
        if spec.representative_genes:
            representative_genes_by_spec[spec] = spec.representative_genes
            continue

        examples: list[str] = []
        seen: set[str] = set()
        for row in sorted_rows:
            if row.gene_symbol in seen:
                continue
            if not _matches(row, spec):
                continue
            seen.add(row.gene_symbol)
            examples.append(row.gene_symbol)
            if len(examples) >= max_genes:
                break

        representative_genes_by_spec[spec] = tuple(examples)

    return representative_genes_by_spec


def load_workbook_rows(
    workbook: Path,
    sheet_name: str = DEFAULT_WORKBOOK_SHEET,
) -> list[PNWorkbookRow]:
    """Load gene-level PN rows from the workbook."""
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    index = {str(name).strip(): i for i, name in enumerate(header)}

    workbook_rows: list[PNWorkbookRow] = []
    for row in rows:
        gene_symbol = _clean_value(row[index["Gene Symbol"]])
        if not gene_symbol:
            continue

        raw_order = row[index["order"]] if "order" in index else None
        order = int(raw_order) if raw_order not in (None, "") else None
        workbook_rows.append(
            PNWorkbookRow(
                order=order,
                gene_symbol=gene_symbol,
                gene_name=_clean_value(row[index["Gene Name"]]),
                branch=_clean_value(row[index[WORKBOOK_COLUMNS["branch"]]]),
                class_name=_clean_value(row[index[WORKBOOK_COLUMNS["class"]]]),
                group=_clean_value(row[index[WORKBOOK_COLUMNS["group"]]]),
                type_name=_clean_value(row[index[WORKBOOK_COLUMNS["type"]]]),
                subtype=_clean_value(row[index[WORKBOOK_COLUMNS["subtype"]]]),
                uniprot_id=_clean_value(row[index["UniProt ID"]]) if "UniProt ID" in index else "",
            )
        )

    return workbook_rows


def load_mapping_specs(mapping_dir: Path) -> list[MappingSpec]:
    """Load mapping entries from YAML mapping-set files."""
    specs: list[MappingSpec] = []
    for path in sorted(mapping_dir.glob("*.yaml")):
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        for entry in data.get("mappings", []):
            target_term = entry.get("target_term", {}) or {}
            specs.append(
                MappingSpec(
                    file_name=path.name,
                    subject_code=_clean_value(entry.get("subject_code")),
                    subject_level=_clean_value(entry.get("subject_level")),
                    target_go_id=_clean_value(target_term.get("id")),
                    target_go_label=_clean_value(target_term.get("label")),
                    mapping_scope=_clean_value(entry.get("mapping_scope")),
                    representative_genes=tuple(
                        _clean_value(gene)
                        for gene in entry.get("representative_genes", [])
                        if _clean_value(gene)
                    ),
                    conditions=_normalize_conditions(entry.get("conditions")),
                    rationale=_clean_value(entry.get("rationale")),
                    notes=_clean_value(entry.get("notes")),
                    references=tuple(
                        _clean_value(reference)
                        for reference in entry.get("references", [])
                        if _clean_value(reference)
                    ),
                )
            )
    return specs


class GOClosureClassifier:
    """Classify projected GO terms against existing GOA using ontology closure."""

    def __init__(self, adapter_string: str = "sqlite:obo:go") -> None:
        from oaklib import get_adapter

        self.adapter = get_adapter(adapter_string)
        self._ancestor_cache: dict[str, set[str]] = {}

    def ancestors(self, term_id: str) -> set[str]:
        if term_id not in self._ancestor_cache:
            ancestors = set()
            try:
                ancestors = {
                    ancestor
                    for ancestor in self.adapter.ancestors(
                        term_id, predicates=list(GO_CLOSURE_PREDICATES)
                    )
                    if str(ancestor).startswith("GO:") and ancestor != term_id
                }
            except Exception:
                ancestors = set()
            self._ancestor_cache[term_id] = ancestors
        return self._ancestor_cache[term_id]

    def classify(
        self,
        projected_go_id: str,
        goa_terms: dict[str, str],
    ) -> tuple[str, tuple[str, ...]]:
        if projected_go_id in goa_terms:
            label = goa_terms[projected_go_id]
            return "already_in_goa_exact", (f"{projected_go_id} {label}",)

        entailing_terms: set[str] = set()
        projected_ancestors = self.ancestors(projected_go_id)
        broader_terms: set[str] = set()

        for goa_id, goa_label in goa_terms.items():
            goa_ancestors = self.ancestors(goa_id)
            if projected_go_id in goa_ancestors:
                entailing_terms.add(f"{goa_id} {goa_label}")
            elif goa_id in projected_ancestors:
                broader_terms.add(f"{goa_id} {goa_label}")

        if entailing_terms:
            return "entailed_by_goa_closure", tuple(sorted(entailing_terms))
        if broader_terms:
            return "more_specific_than_existing_goa", tuple(sorted(broader_terms))
        return "new_to_goa", ()


def project_annotations(
    workbook_rows: list[PNWorkbookRow],
    mapping_specs: list[MappingSpec],
    goa_root: Path,
    fetch_missing_goa: bool = False,
    goa_cache_dir: Path | None = None,
) -> list[ProjectedAnnotation]:
    """Project GO annotations onto PN gene rows via curated mappings."""
    validator = GOAValidator()
    closure_classifier = GOClosureClassifier()
    goa_cache: dict[str, tuple[bool, dict[str, str]]] = {}
    projected: list[ProjectedAnnotation] = []
    representative_genes_by_spec = derive_representative_genes(workbook_rows, mapping_specs)
    if goa_cache_dir is None:
        goa_cache_dir = DEFAULT_GOA_CACHE_DIR

    def load_cached_or_remote_goa(row: PNWorkbookRow) -> tuple[bool, dict[str, str]]:
        if not fetch_missing_goa or not row.uniprot_id:
            return False, {}

        goa_cache_dir.mkdir(parents=True, exist_ok=True)
        cache_path = goa_cache_dir / f"{row.gene_symbol}__{row.uniprot_id}-goa.tsv"
        try:
            if not cache_path.exists():
                goa_data = fetch_goa_data(row.uniprot_id)
                cache_path.write_text(goa_data, encoding="utf-8")

            terms = {
                annotation.go_id: annotation.go_term
                for annotation in validator.parse_goa_file(cache_path)
            }
            return True, terms
        except Exception:
            return False, {}

    def load_goa_terms(row: PNWorkbookRow) -> tuple[bool, dict[str, str]]:
        gene_symbol = row.gene_symbol
        if gene_symbol not in goa_cache:
            goa_file = goa_root / gene_symbol / f"{gene_symbol}-goa.tsv"
            if goa_file.exists():
                terms = {
                    annotation.go_id: annotation.go_term
                    for annotation in validator.parse_goa_file(goa_file)
                }
                goa_cache[gene_symbol] = (True, terms)
            else:
                goa_cache[gene_symbol] = load_cached_or_remote_goa(row)
        return goa_cache[gene_symbol]

    for row in workbook_rows:
        has_goa_file, goa_terms = load_goa_terms(row)
        pn_code = _deepest_code(row)

        for spec in mapping_specs:
            if not _matches(row, spec):
                continue

            if not has_goa_file:
                goa_status = "no_local_goa"
                supporting_goa_terms: tuple[str, ...] = ()
            else:
                goa_status, supporting_goa_terms = closure_classifier.classify(
                    spec.target_go_id, goa_terms
                )

            projected.append(
                ProjectedAnnotation(
                    gene_symbol=row.gene_symbol,
                    gene_name=row.gene_name,
                    pn_code=pn_code,
                    branch=row.branch,
                    class_name=row.class_name,
                    group=row.group,
                    type_name=row.type_name,
                    subtype=row.subtype,
                    mapping_file=spec.file_name,
                    mapping_subject_level=spec.subject_level,
                    mapping_subject_code=spec.subject_code,
                    mapping_scope=spec.mapping_scope,
                    target_go_id=spec.target_go_id,
                    target_go_label=spec.target_go_label,
                    representative_genes=representative_genes_by_spec[spec],
                    conditions=spec.conditions,
                    rationale=spec.rationale,
                    notes=spec.notes,
                    references=spec.references,
                    goa_status=goa_status,
                    supporting_goa_terms=supporting_goa_terms,
                )
            )

    return projected


def summarize_gene_go_projections(
    projected_rows: list[ProjectedAnnotation],
) -> list[dict[str, str]]:
    """Collapse row-level projections to unique gene-GO pairs."""
    grouped: dict[tuple[str, str], list[ProjectedAnnotation]] = {}
    for row in projected_rows:
        grouped.setdefault((row.gene_symbol, row.target_go_id), []).append(row)

    summaries: list[dict[str, str]] = []
    for (gene_symbol, go_id), rows in sorted(grouped.items()):
        first = rows[0]
        summaries.append(
            {
                "gene_symbol": gene_symbol,
                "gene_name": first.gene_name,
                "target_go_id": go_id,
                "target_go_label": first.target_go_label,
                "goa_status": first.goa_status,
                "mapping_scopes": ";".join(sorted({row.mapping_scope for row in rows})),
                "mapping_files": ";".join(sorted({row.mapping_file for row in rows})),
                "mapping_subjects": ";".join(sorted({row.mapping_subject_code for row in rows})),
                "pn_codes": ";".join(sorted({row.pn_code for row in rows})),
                "representative_genes": ";".join(
                    sorted({gene for row in rows for gene in row.representative_genes})
                ),
                "projection_count": str(len(rows)),
                "supporting_goa_terms": ";".join(
                    sorted({term for row in rows for term in row.supporting_goa_terms})
                ),
            }
        )

    return summaries


def _format_conditions(conditions: tuple[tuple[str, str], ...]) -> str:
    return "; ".join(f"{level}={code}" for level, code in conditions)


def _format_references(references: tuple[str, ...]) -> str:
    return "; ".join(references)


def write_projection_reports(
    output_dir: Path,
    projected_rows: list[ProjectedAnnotation],
    summary_rows: list[dict[str, str]],
) -> None:
    """Write project-local TSV reports for PN-driven GO projections."""
    output_dir.mkdir(parents=True, exist_ok=True)

    row_level_path = output_dir / "pn_projected_annotations.tsv"
    with row_level_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(
            [
                "gene_symbol",
                "gene_name",
                "pn_code",
                "branch",
                "class",
                "group",
                "type",
                "subtype",
                "mapping_file",
                "mapping_subject_level",
                "mapping_subject_code",
                "mapping_scope",
                "target_go_id",
                "target_go_label",
                "representative_genes",
                "conditions",
                "goa_status",
                "supporting_goa_terms",
                "rationale",
                "notes",
                "references",
            ]
        )
        for row in projected_rows:
            writer.writerow(
                [
                    row.gene_symbol,
                    row.gene_name,
                    row.pn_code,
                    row.branch,
                    row.class_name,
                    row.group,
                    row.type_name,
                    row.subtype,
                    row.mapping_file,
                    row.mapping_subject_level,
                    row.mapping_subject_code,
                    row.mapping_scope,
                    row.target_go_id,
                    row.target_go_label,
                    ";".join(row.representative_genes),
                    _format_conditions(row.conditions),
                    row.goa_status,
                    ";".join(row.supporting_goa_terms),
                    row.rationale,
                    row.notes,
                    _format_references(row.references),
                ]
            )

    summary_path = output_dir / "pn_projected_gene_go_summary.tsv"
    with summary_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            delimiter="\t",
            fieldnames=[
                "gene_symbol",
                "gene_name",
                "target_go_id",
                "target_go_label",
                "goa_status",
                "mapping_scopes",
                "mapping_files",
                "mapping_subjects",
                "pn_codes",
                "representative_genes",
                "projection_count",
                "supporting_goa_terms",
            ],
        )
        writer.writeheader()
        writer.writerows(summary_rows)

    new_to_goa_path = output_dir / "pn_projected_new_to_goa.tsv"
    with new_to_goa_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            delimiter="\t",
            fieldnames=[
                "gene_symbol",
                "gene_name",
                "target_go_id",
                "target_go_label",
                "goa_status",
                "mapping_scopes",
                "mapping_files",
                "mapping_subjects",
                "pn_codes",
                "representative_genes",
                "projection_count",
                "supporting_goa_terms",
            ],
        )
        writer.writeheader()
        writer.writerows(row for row in summary_rows if row["goa_status"] == "new_to_goa")

    candidate_additions_path = output_dir / "pn_projected_candidate_additions.tsv"
    with candidate_additions_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            delimiter="\t",
            fieldnames=[
                "gene_symbol",
                "gene_name",
                "target_go_id",
                "target_go_label",
                "goa_status",
                "mapping_scopes",
                "mapping_files",
                "mapping_subjects",
                "pn_codes",
                "representative_genes",
                "projection_count",
                "supporting_goa_terms",
            ],
        )
        writer.writeheader()
        writer.writerows(
            row
            for row in summary_rows
            if row["goa_status"] in {"new_to_goa", "more_specific_than_existing_goa"}
        )

    status_counts: dict[str, int] = {}
    for row in summary_rows:
        status_counts[row["goa_status"]] = status_counts.get(row["goa_status"], 0) + 1
    status_summary_path = output_dir / "pn_projected_status_summary.tsv"
    with status_summary_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(["metric", "value"])
        writer.writerow(["row_level_projections", str(len(projected_rows))])
        writer.writerow(["unique_gene_go_pairs", str(len(summary_rows))])
        writer.writerow(
            [
                "candidate_additions",
                str(
                    sum(
                        1
                        for row in summary_rows
                        if row["goa_status"] in {"new_to_goa", "more_specific_than_existing_goa"}
                    )
                ),
            ]
        )
        for status in (
            "already_in_goa_exact",
            "entailed_by_goa_closure",
            "more_specific_than_existing_goa",
            "new_to_goa",
            "no_local_goa",
        ):
            writer.writerow([status, str(status_counts.get(status, 0))])


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Project GO annotations from PN workbook rows via curated PN-to-GO mappings."
    )
    parser.add_argument("--workbook", type=Path, required=True, help="PN workbook XLSX path")
    parser.add_argument(
        "--sheet",
        default=DEFAULT_WORKBOOK_SHEET,
        help=f"Workbook sheet to read (default: {DEFAULT_WORKBOOK_SHEET})",
    )
    parser.add_argument(
        "--mapping-dir",
        type=Path,
        default=DEFAULT_MAPPING_DIR,
        help=f"Directory containing mapping YAMLs (default: {DEFAULT_MAPPING_DIR})",
    )
    parser.add_argument(
        "--goa-root",
        type=Path,
        default=DEFAULT_GOA_ROOT,
        help=f"Root directory containing per-gene GOA folders (default: {DEFAULT_GOA_ROOT})",
    )
    parser.add_argument(
        "--fetch-missing-goa",
        action="store_true",
        help="Fetch missing GOA on the fly into a project-local cache instead of requiring gene folders.",
    )
    parser.add_argument(
        "--goa-cache-dir",
        type=Path,
        default=DEFAULT_GOA_CACHE_DIR,
        help=f"Project-local GOA cache directory for on-the-fly fetches (default: {DEFAULT_GOA_CACHE_DIR})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Output report directory (default: {DEFAULT_OUTPUT_DIR})",
    )
    return parser


def main() -> None:
    parser = build_argument_parser()
    args = parser.parse_args()

    workbook_rows = load_workbook_rows(args.workbook, sheet_name=args.sheet)
    mapping_specs = load_mapping_specs(args.mapping_dir)
    projected_rows = project_annotations(
        workbook_rows,
        mapping_specs,
        args.goa_root,
        fetch_missing_goa=args.fetch_missing_goa,
        goa_cache_dir=args.goa_cache_dir,
    )
    summary_rows = summarize_gene_go_projections(projected_rows)
    write_projection_reports(args.output_dir, projected_rows, summary_rows)

    status_counts: dict[str, int] = {}
    for row in summary_rows:
        status_counts[row["goa_status"]] = status_counts.get(row["goa_status"], 0) + 1

    print(
        f"Wrote PN projection reports to {args.output_dir} "
        f"(row_level={len(projected_rows)} unique_gene_go={len(summary_rows)} "
        f"already_in_goa_exact={status_counts.get('already_in_goa_exact', 0)} "
        f"entailed_by_goa_closure={status_counts.get('entailed_by_goa_closure', 0)} "
        f"more_specific_than_existing_goa={status_counts.get('more_specific_than_existing_goa', 0)} "
        f"new_to_goa={status_counts.get('new_to_goa', 0)} "
        f"no_local_goa={status_counts.get('no_local_goa', 0)})"
    )


if __name__ == "__main__":
    main()
