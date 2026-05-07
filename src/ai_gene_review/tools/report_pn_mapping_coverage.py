"""Report mapping coverage for Proteostasis Network source codes.

This tool compares canonical PN codes extracted from the workbook against one
curator-facing record per PN node in the mapping-set YAML files. It distinguishes
among:

- pending_review: accounted for in YAML, but not yet manually analyzed in depth
- mapped: reviewed and mapped to GO
- context_only: related GO context recorded, but unsafe to propagate
- no_mapping: reviewed and concluded no GO mapping is appropriate
- deferred: reviewed but deferred pending evidence, a better term, or narrower handling
- missing_from_yaml: present in the workbook but absent from curation YAML

The coverage report is code-centric rather than gene-centric. Codes are
canonicalized as hierarchy-aware path strings:

- branch: ``Branch``
- class: ``Branch|Class``
- group: ``Branch|Class|Group``
- type: ``Branch|Class|Group|Type``
- subtype: ``Branch|Class|Group|Type|Subtype``
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import yaml

DEFAULT_WORKBOOK_SHEET = "dense"
DEFAULT_MAPPING_DIR = Path("projects/PROTEOSTASIS/mappings")
DEFAULT_OUTPUT_DIR = Path("projects/PROTEOSTASIS/reports/pn_mapping_coverage")
LEVELS = ("branch", "class", "group", "type", "subtype")
WORKBOOK_COLUMNS = {
    "branch": "Branch",
    "class": "Class",
    "group": "Group",
    "type": "Type",
    "subtype": "Subtype",
}
MISSING_MARKERS = {"", "(no Branch)", "(no Class)", "(no Group)", "(no Type)", "(no Subtype)"}
COVERAGE_STATUSES = (
    "pending_review",
    "mapped",
    "context_only",
    "no_mapping",
    "deferred",
    "missing_from_yaml",
    "curation_conflict",
)


@dataclass(frozen=True)
class PNCodeRecord:
    """A canonical PN code derived from the workbook."""

    level: str
    code: str
    branch: str
    class_name: str
    group: str
    type_name: str
    subtype: str


@dataclass(frozen=True)
class SubjectSpec:
    """A source-node curation record from a mapping-set file."""

    file_name: str
    subject_code: str
    subject_level: str
    curation_status: str
    mapping_scope: str
    conditions: tuple[tuple[str, str], ...]


def _clean_value(value: Any) -> str:
    """Normalize workbook or YAML values to a simple stripped string."""
    if value is None:
        return ""
    cleaned = str(value).strip()
    if cleaned in MISSING_MARKERS:
        return ""
    return cleaned


def _canonical_code(level_values: dict[str, str], level: str) -> str:
    """Build a canonical path-like code string through the requested level."""
    parts: list[str] = []
    for current_level in LEVELS:
        value = level_values[current_level]
        if not value:
            break
        parts.append(value)
        if current_level == level:
            break
    return "|".join(parts)


def load_workbook_codes(
    workbook: Path,
    sheet_name: str = DEFAULT_WORKBOOK_SHEET,
) -> list[PNCodeRecord]:
    """Extract unique canonical codes at each hierarchy level from the workbook."""
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    index = {str(name).strip(): i for i, name in enumerate(header)}

    unique: dict[tuple[str, str], PNCodeRecord] = {}
    for row in rows:
        if not _clean_value(row[index["Gene Symbol"]]):
            continue

        level_values = {
            level: _clean_value(row[index[column]])
            for level, column in WORKBOOK_COLUMNS.items()
        }

        for level in LEVELS:
            if not level_values[level]:
                continue
            code = _canonical_code(level_values, level)
            key = (level, code)
            unique.setdefault(
                key,
                PNCodeRecord(
                    level=level,
                    code=code,
                    branch=level_values["branch"],
                    class_name=level_values["class"],
                    group=level_values["group"],
                    type_name=level_values["type"],
                    subtype=level_values["subtype"],
                ),
            )

    return sorted(unique.values(), key=lambda record: (LEVELS.index(record.level), record.code))


def _normalize_conditions(raw_conditions: Any) -> tuple[tuple[str, str], ...]:
    """Convert a YAML conditions list into a canonical tuple form."""
    if not raw_conditions:
        return ()

    conditions: list[tuple[str, str]] = []
    for condition in raw_conditions:
        conditions.append(
            (
                _clean_value(condition.get("condition_level")),
                _clean_value(condition.get("condition_code")),
            )
        )
    return tuple(conditions)


def load_subject_specs(mapping_dir: Path) -> list[SubjectSpec]:
    """Load source-node curation specs from mapping-set files."""
    specs: list[SubjectSpec] = []
    for path in sorted(mapping_dir.glob("*.yaml")):
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        for entry in data.get("subject_curations", []):
            specs.append(
                SubjectSpec(
                    file_name=path.name,
                    subject_code=_clean_value(entry.get("subject_code")),
                    subject_level=_clean_value(entry.get("subject_level")),
                    curation_status=_clean_value(entry.get("curation_status")),
                    mapping_scope=_clean_value(entry.get("mapping_scope")),
                    conditions=_normalize_conditions(entry.get("conditions")),
                )
            )
    return specs


def _record_level_value(record: PNCodeRecord, level: str) -> str:
    """Return the plain value at a given level for a code record."""
    if level == "branch":
        return record.branch
    if level == "class":
        return record.class_name
    if level == "group":
        return record.group
    if level == "type":
        return record.type_name
    if level == "subtype":
        return record.subtype
    raise ValueError(f"Unsupported level: {level}")


def _record_code_at_level(record: PNCodeRecord, level: str) -> str:
    """Return the canonical code prefix for a record at a given level."""
    return _canonical_code(
        {
            "branch": record.branch,
            "class": record.class_name,
            "group": record.group,
            "type": record.type_name,
            "subtype": record.subtype,
        },
        level,
    )


def _subject_matches_record(record: PNCodeRecord, spec: SubjectSpec) -> bool:
    """Return True if a subject spec covers this canonical code record."""
    if record.level != spec.subject_level:
        return False

    if "|" in spec.subject_code:
        if _record_code_at_level(record, spec.subject_level) != spec.subject_code:
            return False
    elif _record_level_value(record, spec.subject_level) != spec.subject_code:
        return False

    for condition_level, condition_code in spec.conditions:
        if "|" in condition_code:
            if _record_code_at_level(record, condition_level) != condition_code:
                return False
        elif _record_level_value(record, condition_level) != condition_code:
            return False

    return True


def summarize_coverage(
    code_records: list[PNCodeRecord],
    curation_specs: list[SubjectSpec],
) -> list[dict[str, str]]:
    """Compute code-level mapping coverage rows."""
    rows: list[dict[str, str]] = []

    for record in code_records:
        matched_curations = [
            spec for spec in curation_specs if _subject_matches_record(record, spec)
        ]

        if not matched_curations:
            status = "missing_from_yaml"
        elif len(matched_curations) > 1:
            status = "curation_conflict"
        else:
            status = matched_curations[0].curation_status

        rows.append(
            {
                "level": record.level,
                "code": record.code,
                "branch": record.branch,
                "class": record.class_name,
                "group": record.group,
                "type": record.type_name,
                "subtype": record.subtype,
                "status": status,
                "curation_files": ";".join(
                    sorted({spec.file_name for spec in matched_curations})
                ),
                "curation_subjects": ";".join(
                    sorted({spec.subject_code for spec in matched_curations})
                ),
                "curation_statuses": ";".join(
                    sorted({spec.curation_status for spec in matched_curations})
                ),
                "mapping_scopes": ";".join(
                    sorted({spec.mapping_scope for spec in matched_curations if spec.mapping_scope})
                ),
            }
        )

    return rows


def write_coverage_outputs(rows: list[dict[str, str]], output_dir: Path) -> None:
    """Write TSV outputs for detailed code coverage and per-level summaries."""
    output_dir.mkdir(parents=True, exist_ok=True)

    detail_path = output_dir / "pn_mapping_code_coverage.tsv"
    detail_fields = [
        "level",
        "code",
        "branch",
        "class",
        "group",
        "type",
        "subtype",
        "status",
        "mapping_scopes",
        "curation_files",
        "curation_subjects",
        "curation_statuses",
    ]
    with detail_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=detail_fields,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)

    summary_path = output_dir / "pn_mapping_level_summary.tsv"
    summary_fields = [
        "level",
        "total_codes",
        "pending_review",
        "mapped",
        "context_only",
        "no_mapping",
        "deferred",
        "missing_from_yaml",
        "curation_conflict",
    ]
    with summary_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=summary_fields,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        for level in LEVELS:
            level_rows = [row for row in rows if row["level"] == level]
            writer.writerow(
                {
                    "level": level,
                    "total_codes": len(level_rows),
                    **{
                        status: sum(row["status"] == status for row in level_rows)
                        for status in COVERAGE_STATUSES
                    },
                }
            )


def build_argument_parser() -> argparse.ArgumentParser:
    """Build the CLI argument parser."""
    parser = argparse.ArgumentParser(
        description="Report coverage of PN source codes by mapping-set YAML files."
    )
    parser.add_argument("--workbook", required=True, type=Path, help="Path to the PN workbook")
    parser.add_argument(
        "--sheet-name",
        default=DEFAULT_WORKBOOK_SHEET,
        help=f"Workbook sheet name to use (default: {DEFAULT_WORKBOOK_SHEET})",
    )
    parser.add_argument(
        "--mapping-dir",
        type=Path,
        default=DEFAULT_MAPPING_DIR,
        help=f"Directory containing mapping-set YAML files (default: {DEFAULT_MAPPING_DIR})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Directory for TSV outputs (default: {DEFAULT_OUTPUT_DIR})",
    )
    return parser


def main() -> None:
    """Run the PN mapping coverage report."""
    parser = build_argument_parser()
    args = parser.parse_args()

    code_records = load_workbook_codes(args.workbook, sheet_name=args.sheet_name)
    curation_specs = load_subject_specs(args.mapping_dir)
    rows = summarize_coverage(code_records, curation_specs)
    write_coverage_outputs(rows, args.output_dir)

    total = len(rows)
    mapped = sum(row["status"] == "mapped" for row in rows)
    no_mapping = sum(row["status"] == "no_mapping" for row in rows)
    deferred = sum(row["status"] == "deferred" for row in rows)
    pending_review = sum(row["status"] == "pending_review" for row in rows)
    context_only = sum(row["status"] == "context_only" for row in rows)
    missing = sum(row["status"] == "missing_from_yaml" for row in rows)
    conflicts = sum(row["status"] == "curation_conflict" for row in rows)

    print(f"Wrote coverage reports to {args.output_dir}")
    print(
        "Coverage summary: "
        f"total={total} pending_review={pending_review} mapped={mapped} "
        f"context_only={context_only} no_mapping={no_mapping} "
        f"deferred={deferred} "
        f"missing_from_yaml={missing} conflicts={conflicts}"
    )


if __name__ == "__main__":
    main()
