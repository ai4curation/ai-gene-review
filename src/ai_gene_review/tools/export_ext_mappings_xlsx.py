"""Export curated external mapping YAML files to a biologist-friendly XLSX."""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import yaml

from ai_gene_review.tools.report_pn_mapping_coverage import LEVELS
from ai_gene_review.tools.report_pn_projected_annotations import (
    DEFAULT_GOA_CACHE_DIR,
    DEFAULT_GOA_DUCKDB,
    DEFAULT_WORKBOOK_SHEET,
    derive_representative_genes,
    load_mapping_specs as load_projection_mapping_specs,
    load_workbook_rows as load_projection_workbook_rows,
    project_annotations,
    summarize_gene_go_projections,
)

DEFAULT_MAPPING_DIR = Path("projects/PROTEOSTASIS/mappings")
DEFAULT_OUTPUT_PATH = Path("projects/PROTEOSTASIS/reports/pn_mappings/pn_mappings.xlsx")
DEFAULT_GOA_ROOT = Path("genes/human")

README_ROWS = [
    ("Proteostasis PN to GO mapping export", ""),
    (
        "Overview",
        "This workbook exports manually curated mappings from Proteostasis Network "
        "(PN) hierarchy codes to Gene Ontology (GO) terms. The source mappings live "
        "in project YAML files and are intended as a curation aid, not as an "
        "authoritative ontology release.",
    ),
    (
        "How to read subject_code",
        "subject_code is a pipe-delimited PN path in the form "
        "Branch|Class|Group|Type|Subtype. The split hierarchy columns in the data "
        "tabs repeat those levels in separate columns for filtering and sorting.",
    ),
    (
        "Mappings tab",
        "One row per GO-bearing PN curation. This includes propagating mappings "
        "and context-only GO relationships.",
    ),
    (
        "Unmapped tab",
        "One row per PN subject without a GO target. curation_status separates "
        "pending review, deferred decisions, and reviewed no-mapping decisions.",
    ),
    (
        "Summary tab",
        "Count summary by mapping YAML file, including the number of GO-bearing "
        "curations and non-mapping curations.",
    ),
    (
        "Per-file tabs",
        "Each additional sheet corresponds to one mapping YAML file and includes both "
        "mapped and unmapped rows for that file.",
    ),
    (
        "Projection tabs",
        "When the export is run with a PN workbook, the workbook also includes "
        "Projected Gene-GO, Candidate Additions, and Projection Status tabs. "
        "These summarize what GO annotations would be induced by the current PN "
        "mappings and how they compare with local GOA. If the export is run with "
        "--fetch-missing-goa, missing GOA is fetched into a project-local cache "
        "instead of requiring per-gene review folders.",
    ),
    (
        "subject_level",
        "The PN level for the subject_code being mapped: branch, class, group, type, "
        "or subtype.",
    ),
    (
        "curation_status",
        "pending_review means the PN node is accounted for but not yet manually "
        "analyzed. mapped means a GO mapping is complete. context_only means a GO "
        "relationship is recorded but unsafe to propagate. no_mapping means reviewed "
        "and no GO mapping should be made. deferred means review is blocked by "
        "evidence, taxonomy ambiguity, or a missing/better GO term.",
    ),
    (
        "mapping_scope",
        "exact means the PN code and GO term are treated as semantically equivalent. "
        "ok_for_propagation_to_go means the PN code is not identical to the GO term, "
        "but genes in that PN bucket can reasonably propagate to the GO target. "
        "too_broad_to_propagate means the PN-to-GO relationship is real and worth "
        "recording, but projecting it onto member genes would over-annotate them.",
    ),
    (
        "target_go_id / target_go_label",
        "The GO term used as the mapping target. The export keeps both ID and label so "
        "ontology validation can check that the pair is internally consistent.",
    ),
    (
        "representative_genes",
        "Optional example genes that illustrate what the PN bucket means in "
        "practice. These are explanatory examples, not a complete member list. "
        "When the export is run with a PN workbook, blank manual fields are "
        "filled from representative workbook members.",
    ),
    (
        "conditions",
        "Optional extra PN context required for the mapping, formatted as "
        "level=code. Blank means the mapping applies without additional conditions.",
    ),
    (
        "rationale",
        "Manual curator explanation of why the mapping is defensible.",
    ),
    (
        "notes",
        "Additional curator commentary, caveats, or scope limitations.",
    ),
    (
        "references",
        "Project-local provenance tags for the workbook, manuscripts, or other curation "
        "sources used to justify the mapping.",
    ),
    (
        "goa_status",
        "Projection status relative to local GOA. already_in_goa_exact means the exact "
        "GO term is already present. entailed_by_goa_closure means local GOA already "
        "contains a more specific descendant. more_specific_than_existing_goa means the "
        "PN projection is more specific than an existing broader GOA term. "
        "supported_by_goa_regulation means GOA already contains regulation of the "
        "projected process, so the projection may still be useful but should not be "
        "treated as wholly novel. new_to_goa means no exact, closure-based, or "
        "regulation-based support was found. no_local_goa means no GOA record was "
        "available for that gene from the configured GOA source during the export run.",
    ),
]


@dataclass(frozen=True)
class MappingRow:
    """Flat exported view of one GO-bearing curation."""

    mapping_file: str
    mapping_set_id: str
    mapping_set_label: str
    subject_level: str
    subject_code: str
    curation_status: str
    branch: str
    class_name: str
    group: str
    type_name: str
    subtype: str
    mapping_scope: str
    target_go_id: str
    target_go_label: str
    representative_genes: str
    conditions: str
    rationale: str
    notes: str
    references: str


@dataclass(frozen=True)
class UnmappedRow:
    """Flat exported view of one non-mapping curation."""

    mapping_file: str
    mapping_set_id: str
    mapping_set_label: str
    subject_level: str
    subject_code: str
    curation_status: str
    branch: str
    class_name: str
    group: str
    type_name: str
    subtype: str
    conditions: str
    rationale: str
    notes: str
    references: str


def _clean_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _split_subject_code(subject_code: str) -> dict[str, str]:
    parts = [_clean_value(part) for part in subject_code.split("|")] if subject_code else []
    padded = parts + [""] * (len(LEVELS) - len(parts))
    return {
        "branch": padded[0],
        "class": padded[1],
        "group": padded[2],
        "type": padded[3],
        "subtype": padded[4],
    }


def _format_conditions(raw_conditions: Any) -> str:
    if not raw_conditions:
        return ""
    parts = []
    for condition in raw_conditions:
        level = _clean_value(condition.get("condition_level"))
        code = _clean_value(condition.get("condition_code"))
        parts.append(f"{level}={code}")
    return "; ".join(parts)


def _format_references(raw_references: Any) -> str:
    if not raw_references:
        return ""
    return "; ".join(_clean_value(reference) for reference in raw_references)


def _format_string_list(raw_values: Any) -> str:
    if not raw_values:
        return ""
    return "; ".join(_clean_value(value) for value in raw_values if _clean_value(value))


def load_mapping_rows(mapping_dir: Path) -> tuple[list[MappingRow], list[UnmappedRow]]:
    """Load all mapping YAML files into flat row objects."""
    mapping_rows: list[MappingRow] = []
    unmapped_rows: list[UnmappedRow] = []

    for path in sorted(mapping_dir.glob("*.yaml")):
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        mapping_set_id = _clean_value(data.get("id"))
        mapping_set_label = _clean_value(data.get("label"))

        for entry in data.get("subject_curations", []):
            curation_status = _clean_value(entry.get("curation_status"))
            split = _split_subject_code(_clean_value(entry.get("subject_code")))
            target_term = entry.get("target_term", {}) or {}
            if target_term and entry.get("mapping_scope"):
                mapping_rows.append(
                    MappingRow(
                        mapping_file=path.name,
                        mapping_set_id=mapping_set_id,
                        mapping_set_label=mapping_set_label,
                        subject_level=_clean_value(entry.get("subject_level")),
                        subject_code=_clean_value(entry.get("subject_code")),
                        curation_status=curation_status,
                        branch=split["branch"],
                        class_name=split["class"],
                        group=split["group"],
                        type_name=split["type"],
                        subtype=split["subtype"],
                        mapping_scope=_clean_value(entry.get("mapping_scope")),
                        target_go_id=_clean_value(target_term.get("id")),
                        target_go_label=_clean_value(target_term.get("label")),
                        representative_genes=_format_string_list(entry.get("representative_genes")),
                        conditions=_format_conditions(entry.get("conditions")),
                        rationale=_clean_value(entry.get("rationale")),
                        notes=_clean_value(entry.get("notes")),
                        references=_format_references(entry.get("references")),
                    )
                )
            else:
                unmapped_rows.append(
                    UnmappedRow(
                        mapping_file=path.name,
                        mapping_set_id=mapping_set_id,
                        mapping_set_label=mapping_set_label,
                        subject_level=_clean_value(entry.get("subject_level")),
                        subject_code=_clean_value(entry.get("subject_code")),
                        curation_status=curation_status,
                        branch=split["branch"],
                        class_name=split["class"],
                        group=split["group"],
                        type_name=split["type"],
                        subtype=split["subtype"],
                        conditions=_format_conditions(entry.get("conditions")),
                        rationale=_clean_value(entry.get("rationale")),
                        notes=_clean_value(entry.get("notes")),
                        references=_format_references(entry.get("references")),
                    )
                )

    return mapping_rows, unmapped_rows


def _format_condition_tuples(conditions: tuple[tuple[str, str], ...]) -> str:
    if not conditions:
        return ""
    return "; ".join(f"{level}={code}" for level, code in conditions)


def build_mapping_row_representative_lookup(
    mapping_dir: Path,
    workbook_path: Path,
    sheet_name: str = DEFAULT_WORKBOOK_SHEET,
) -> dict[tuple[str, str, str, str, str], str]:
    """Derive example genes for mapping rows from matched workbook members."""
    workbook_rows = load_projection_workbook_rows(workbook_path, sheet_name=sheet_name)
    mapping_specs = load_projection_mapping_specs(mapping_dir)
    representative_genes_by_spec = derive_representative_genes(workbook_rows, mapping_specs)

    lookup: dict[tuple[str, str, str, str, str], str] = {}
    for spec in mapping_specs:
        lookup[
            (
                spec.file_name,
                spec.subject_level,
                spec.subject_code,
                spec.target_go_id,
                _format_condition_tuples(spec.conditions),
            )
        ] = "; ".join(representative_genes_by_spec[spec])
    return lookup


def apply_derived_representative_genes(
    mapping_rows: list[MappingRow],
    representative_lookup: dict[tuple[str, str, str, str, str], str],
) -> list[MappingRow]:
    """Fill blank representative_genes cells from derived workbook exemplars."""
    hydrated_rows: list[MappingRow] = []
    for row in mapping_rows:
        if row.representative_genes:
            hydrated_rows.append(row)
            continue

        key = (
            row.mapping_file,
            row.subject_level,
            row.subject_code,
            row.target_go_id,
            row.conditions,
        )
        hydrated_rows.append(
            replace(
                row,
                representative_genes=representative_lookup.get(key, row.representative_genes),
            )
        )
    return hydrated_rows


def build_projection_tables(
    mapping_dir: Path,
    workbook_path: Path,
    goa_root: Path,
    sheet_name: str = DEFAULT_WORKBOOK_SHEET,
    goa_duckdb: Path | None = None,
    fetch_missing_goa: bool = False,
    goa_cache_dir: Path | None = None,
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[list[str]]]:
    """Build projection-derived tables for embedding in the XLSX export."""
    workbook_rows = load_projection_workbook_rows(workbook_path, sheet_name=sheet_name)
    mapping_specs = load_projection_mapping_specs(mapping_dir)
    projected_rows = project_annotations(
        workbook_rows,
        mapping_specs,
        goa_root,
        goa_duckdb=goa_duckdb,
        fetch_missing_goa=fetch_missing_goa,
        goa_cache_dir=goa_cache_dir,
    )
    summary_rows = summarize_gene_go_projections(projected_rows)
    candidate_rows = [
        row
        for row in summary_rows
        if row["goa_status"]
        in {
            "new_to_goa",
            "more_specific_than_existing_goa",
            "supported_by_goa_regulation",
        }
    ]

    status_counts = Counter(row["goa_status"] for row in summary_rows)
    status_rows = [
        ["row_level_projections", str(len(projected_rows))],
        ["unique_gene_go_pairs", str(len(summary_rows))],
        ["candidate_additions", str(len(candidate_rows))],
        ["already_in_goa_exact", str(status_counts.get("already_in_goa_exact", 0))],
        ["entailed_by_goa_closure", str(status_counts.get("entailed_by_goa_closure", 0))],
        [
            "more_specific_than_existing_goa",
            str(status_counts.get("more_specific_than_existing_goa", 0)),
        ],
        [
            "supported_by_goa_regulation",
            str(status_counts.get("supported_by_goa_regulation", 0)),
        ],
        ["new_to_goa", str(status_counts.get("new_to_goa", 0))],
        ["no_local_goa", str(status_counts.get("no_local_goa", 0))],
    ]
    return summary_rows, candidate_rows, status_rows


def _apply_table_formatting(worksheet) -> None:
    """Apply lightweight spreadsheet formatting for readability."""
    if worksheet.max_row == 0 or worksheet.max_column == 0:
        return

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    width_overrides = {
        "A": 24,
        "B": 26,
        "C": 28,
        "D": 12,
        "E": 50,
        "F": 24,
        "G": 28,
        "H": 28,
        "I": 28,
        "J": 28,
        "K": 18,
        "L": 14,
        "M": 34,
        "N": 28,
        "O": 28,
        "P": 60,
        "Q": 48,
        "R": 36,
    }

    for column_idx in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column_idx)
        worksheet.column_dimensions[letter].width = width_overrides.get(letter, 20)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _append_sheet(worksheet, headers: list[str], rows: list[list[str]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    _apply_table_formatting(worksheet)


def _append_readme_sheet(worksheet) -> None:
    """Populate a non-tabular README sheet."""
    for key, value in README_ROWS:
        worksheet.append([key, value])

    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A1"].alignment = Alignment(vertical="top", wrap_text=True)

    for row in worksheet.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[0].alignment = Alignment(vertical="top", wrap_text=True)
        row[1].alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 110
    worksheet.freeze_panes = "A2"


def _sheet_name_for_mapping_file(file_name: str) -> str:
    stem = Path(file_name).stem
    label = stem.replace("_", " ").title()
    return label[:31]


def build_workbook(
    mapping_rows: list[MappingRow],
    unmapped_rows: list[UnmappedRow],
    projected_summary_rows: list[dict[str, str]] | None = None,
    candidate_rows: list[dict[str, str]] | None = None,
    projection_status_rows: list[list[str]] | None = None,
) -> Workbook:
    """Build the export workbook."""
    workbook = Workbook()
    readme_sheet = workbook.active
    readme_sheet.title = "README"
    _append_readme_sheet(readme_sheet)

    summary_sheet = workbook.create_sheet("Summary")
    summary_headers = [
        "mapping_file",
        "mapping_set_id",
        "mapping_set_label",
        "mapping_count",
        "unmapped_count",
    ]
    mapping_counts = Counter(row.mapping_file for row in mapping_rows)
    unmapped_counts = Counter(row.mapping_file for row in unmapped_rows)
    summary_rows = []
    seen_files = sorted(set(mapping_counts) | set(unmapped_counts))
    set_lookup = {
        row.mapping_file: (row.mapping_set_id, row.mapping_set_label) for row in mapping_rows
    }
    for row in unmapped_rows:
        set_lookup.setdefault(row.mapping_file, (row.mapping_set_id, row.mapping_set_label))
    for file_name in seen_files:
        mapping_set_id, mapping_set_label = set_lookup[file_name]
        summary_rows.append(
            [
                file_name,
                mapping_set_id,
                mapping_set_label,
                str(mapping_counts[file_name]),
                str(unmapped_counts[file_name]),
            ]
        )
    _append_sheet(summary_sheet, summary_headers, summary_rows)

    mapping_headers = [
        "mapping_file",
        "mapping_set_id",
        "mapping_set_label",
        "subject_level",
        "subject_code",
        "curation_status",
        "branch",
        "class",
        "group",
        "type",
        "subtype",
        "mapping_scope",
        "target_go_id",
        "target_go_label",
        "representative_genes",
        "conditions",
        "rationale",
        "notes",
        "references",
    ]
    all_mapping_rows = [
        [
            row.mapping_file,
            row.mapping_set_id,
            row.mapping_set_label,
            row.subject_level,
            row.subject_code,
            row.curation_status,
            row.branch,
            row.class_name,
            row.group,
            row.type_name,
            row.subtype,
            row.mapping_scope,
            row.target_go_id,
            row.target_go_label,
            row.representative_genes,
            row.conditions,
            row.rationale,
            row.notes,
            row.references,
        ]
        for row in mapping_rows
    ]
    mappings_sheet = workbook.create_sheet("Mappings")
    _append_sheet(mappings_sheet, mapping_headers, all_mapping_rows)

    unmapped_headers = [
        "mapping_file",
        "mapping_set_id",
        "mapping_set_label",
        "subject_level",
        "subject_code",
        "curation_status",
        "branch",
        "class",
        "group",
        "type",
        "subtype",
        "conditions",
        "rationale",
        "notes",
        "references",
    ]
    all_unmapped_rows = [
        [
            row.mapping_file,
            row.mapping_set_id,
            row.mapping_set_label,
            row.subject_level,
            row.subject_code,
            row.curation_status,
            row.branch,
            row.class_name,
            row.group,
            row.type_name,
            row.subtype,
            row.conditions,
            row.rationale,
            row.notes,
            row.references,
        ]
        for row in unmapped_rows
    ]
    unmapped_sheet = workbook.create_sheet("Unmapped")
    _append_sheet(unmapped_sheet, unmapped_headers, all_unmapped_rows)

    if projected_summary_rows is not None and candidate_rows is not None and projection_status_rows is not None:
        projection_headers = [
            "gene_symbol",
            "gene_name",
            "target_go_id",
            "target_go_label",
            "goa_status",
            "supporting_goa_terms",
            "mapping_scopes",
            "mapping_files",
            "mapping_subjects",
            "pn_codes",
            "representative_genes",
            "projection_count",
        ]
        projected_sheet = workbook.create_sheet("Projected Gene-GO")
        _append_sheet(
            projected_sheet,
            projection_headers,
            [[row.get(header, "") for header in projection_headers] for row in projected_summary_rows],
        )

        candidate_sheet = workbook.create_sheet("Candidate Additions")
        _append_sheet(
            candidate_sheet,
            projection_headers,
            [[row.get(header, "") for header in projection_headers] for row in candidate_rows],
        )

        projection_status_sheet = workbook.create_sheet("Projection Status")
        _append_sheet(projection_status_sheet, ["metric", "value"], projection_status_rows)

    by_file_rows: dict[str, list[list[str]]] = {}
    for row in mapping_rows:
        by_file_rows.setdefault(row.mapping_file, []).append(
            [
                row.curation_status,
                row.subject_level,
                row.subject_code,
                row.branch,
                row.class_name,
                row.group,
                row.type_name,
                row.subtype,
                row.mapping_scope,
                row.curation_status,
                row.target_go_id,
                row.target_go_label,
                row.representative_genes,
                row.conditions,
                row.rationale,
                row.notes,
                row.references,
            ]
        )
    for row in unmapped_rows:
        by_file_rows.setdefault(row.mapping_file, []).append(
            [
                row.curation_status,
                row.subject_level,
                row.subject_code,
                row.branch,
                row.class_name,
                row.group,
                row.type_name,
                row.subtype,
                "",
                row.curation_status,
                "",
                "",
                "",
                row.conditions,
                row.rationale,
                row.notes,
                row.references,
            ]
        )

    per_file_headers = [
        "status",
        "subject_level",
        "subject_code",
        "branch",
        "class",
        "group",
        "type",
        "subtype",
        "mapping_scope",
        "curation_status",
        "target_go_id",
        "target_go_label",
        "representative_genes",
        "conditions",
        "rationale",
        "notes",
        "references",
    ]
    for file_name in seen_files:
        sheet = workbook.create_sheet(_sheet_name_for_mapping_file(file_name))
        _append_sheet(sheet, per_file_headers, by_file_rows.get(file_name, []))

    return workbook


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Export external mapping YAML files to a biologist-friendly XLSX workbook."
    )
    parser.add_argument(
        "--mapping-dir",
        type=Path,
        default=DEFAULT_MAPPING_DIR,
        help=f"Directory containing mapping-set YAML files (default: {DEFAULT_MAPPING_DIR})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"Output XLSX path (default: {DEFAULT_OUTPUT_PATH})",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        help="Optional PN workbook XLSX path. When provided, projection tabs are embedded.",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_WORKBOOK_SHEET,
        help=f"Workbook sheet to read when --workbook is provided (default: {DEFAULT_WORKBOOK_SHEET})",
    )
    parser.add_argument(
        "--goa-root",
        type=Path,
        default=DEFAULT_GOA_ROOT,
        help=f"Root directory containing local GOA gene folders (default: {DEFAULT_GOA_ROOT})",
    )
    parser.add_argument(
        "--goa-duckdb",
        type=Path,
        default=None,
        help=(
            "Optional DuckDB GOA database to use as the primary GOA source, "
            f"for example {DEFAULT_GOA_DUCKDB}"
        ),
    )
    parser.add_argument(
        "--fetch-missing-goa",
        action="store_true",
        help="Fetch missing GOA on the fly into a project-local cache instead of requiring gene folders.",
    )
    parser.add_argument(
        "--goa-cache-dir",
        type=Path,
        default=DEFAULT_GOA_CACHE_DIR,
        help=f"Project-local GOA cache directory for on-the-fly fetches (default: {DEFAULT_GOA_CACHE_DIR})",
    )
    return parser


def main() -> None:
    parser = build_argument_parser()
    args = parser.parse_args()

    mapping_rows, unmapped_rows = load_mapping_rows(args.mapping_dir)
    projected_summary_rows = None
    candidate_rows = None
    projection_status_rows = None
    if args.workbook:
        representative_lookup = build_mapping_row_representative_lookup(
            args.mapping_dir,
            args.workbook,
            sheet_name=args.sheet,
        )
        mapping_rows = apply_derived_representative_genes(mapping_rows, representative_lookup)
        projected_summary_rows, candidate_rows, projection_status_rows = build_projection_tables(
            args.mapping_dir,
            args.workbook,
            args.goa_root,
            sheet_name=args.sheet,
            goa_duckdb=args.goa_duckdb,
            fetch_missing_goa=args.fetch_missing_goa,
            goa_cache_dir=args.goa_cache_dir,
        )

    workbook = build_workbook(
        mapping_rows,
        unmapped_rows,
        projected_summary_rows=projected_summary_rows,
        candidate_rows=candidate_rows,
        projection_status_rows=projection_status_rows,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(
        f"Wrote {args.output} "
        f"(mappings={len(mapping_rows)} unmapped={len(unmapped_rows)})"
    )


if __name__ == "__main__":
    main()
